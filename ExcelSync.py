# -*-coding:utf-8 -*-

import sys,os
from PyQt5.QtWidgets import *

from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
from openpyxl.utils import get_column_letter,column_index_from_string
import xlrd

# 主窗口的设计
class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow,self).__init__(parent)
        self.resize(1000, 600)                                  # 设置窗口大小
        self.status = self.statusBar()                          # 创建状态栏
        self.status.showMessage("this is a status tip",5000)    # 显示状态栏
        self.setWindowTitle("ExcelSync")                        # 窗口标题
        self.ctrWindow = Widget()                               # 定义中心窗口
        self.setCentralWidget(self.ctrWindow)                   # 设置中心窗口


# 主窗口的设计
class Widget(QWidget):
    def __init__(self, parent=None):
        super(Widget,self).__init__(parent)
        self.resize(1000, 600)                                  # 设置窗口大小
        # self.setWindowTitle("ExcelSync")
        self.setupUi()

    def setupUi(self):
        mainLayout = QGridLayout()                              # 主布局
        # 布局1
        hLayout = QHBoxLayout()
        self.comboBox1 = QComboBox()  # 组件
        self.comboBox1.setEditable(True)
        hLayout.addWidget(self.comboBox1, 3, Qt.AlignTop)
        self.btn1 = QPushButton(u"浏览")  # 组件
        self.btn1.clicked.connect(lambda: self.button_respone(self.comboBox1))
        hLayout.addWidget(self.btn1, 1, Qt.AlignTop)
        self.btn2 = QPushButton(u"开始比较")  # 组件
        self.btn2.clicked.connect(lambda: self.compare_start())
        hLayout.addWidget(self.btn2, 1, Qt.AlignTop)
        self.comboBox2 = QComboBox()  # 组件
        self.comboBox2.setEditable(True)
        hLayout.addWidget(self.comboBox2, 3, Qt.AlignTop)
        self.btn3 = QPushButton(u"浏览")  # 组件
        self.btn3.clicked.connect(lambda: self.button_respone(self.comboBox2))
        hLayout.addWidget(self.btn3, 1, Qt.AlignTop)
        mainLayout.addLayout(hLayout, 1, 0)

        # 布局2
        hLayoutExcel = QHBoxLayout()
        self.tabWidgetExcel1 = QTabWidget()
        hLayoutExcel.addWidget(self.tabWidgetExcel1)
        mainLayout.addLayout(hLayoutExcel,2,0)

        # 布局3
        self.hLayout3 = QHBoxLayout()
        self.tabWidgetDiff = QTabWidget()
        self.hLayout3.addWidget(self.tabWidgetDiff)
        mainLayout.addLayout(self.hLayout3, 3, 0)

        self.setLayout(mainLayout)

        self.tabWidgetExcel1.currentChanged.connect(self.update_tab)
        self.tabWidgetDiff.currentChanged.connect(self.update_tab)

        self.last_selected = []

    def button_respone(self, comboBox):
        fileName, filetype = QFileDialog.getOpenFileName(self,
                                                         u"选取Excel文件",
                                                         "./",
                                                         "All Files (*);;Excel Files (*.xlsx;*.xls)")  # 设置文件扩展名过滤,注意用双分号间隔
        if fileName is not None:
            filename_suffix = os.path.splitext(fileName)[-1]
            if filename_suffix.find('.xls') != -1:                         # 处理Excel类型文件
                # 如果存在，先删除，后插入至第一个位置
                # 如果不存在，直接插入
                counts = comboBox.count()
                if counts != 0:
                    for idx in range(counts):
                        text = comboBox.itemText(idx)
                        if text == fileName:
                            comboBox.removeItem(idx)
                comboBox.insertItem(0, fileName)
                comboBox.setCurrentIndex(0)
            else:                                                             # 提示选择Excel类型文件
                QMessageBox.information(self, "warning", u"请选择Excel类型文件")

    def compare_start(self):
        # 处理 是否有文件需要比较
        if self.comboBox1.count() == 0 or self.comboBox2.count() == 0:
            QMessageBox.information(self, "warning", u"请先选择需要比较的Excel文件")
            return
        path1 = self.comboBox1.currentText()
        path2 = self.comboBox2.currentText()

        if not os.path.exists(path1):
            QMessageBox.information(self, "warning", u"文件1或路径不存在")
            return
        if not os.path.exists(path1):
            QMessageBox.information(self, "warning", u"文件2或路径不存在")
            return

        if path1 == path2:
            QMessageBox.information(self, "warning", u"请选择不同文件进行比较")
            return
        # 有文件，处理两个文件中是否有sheet名字一样的情况
        # 防止每次点击开始比较时,出现多次字窗口
        self.tabWidgetExcel1.clear()
        self.tabWidgetDiff.clear()
        self.show_two_excel()

    def show_origin_excel_sheet(self, sheet, tab, max_nrows, max_ncols):
        nrows = sheet.nrows
        ncols = sheet.ncols
        tab.setRowCount(max_nrows)  # 设置表的行数
        tab.setColumnCount(max_ncols)  # 设置表的列数
        table_head = num_converted_into_letters(max_ncols)  # 将数字序列转化成字母序列
        tab.setHorizontalHeaderLabels(table_head)  # 设置表头
        for row in range(nrows):
            for col in range(ncols):
                text = self.to_str(sheet.cell_value(row, col))
                newItem = QTableWidgetItem(text)
                tab.setItem(row, col, newItem)
        tab.update()

    def show_two_excel(self):
        with xlrd.open_workbook(self.comboBox1.currentText()) as workbook1:
            sheet_names_excel1 = workbook1.sheet_names()
            with xlrd.open_workbook(self.comboBox2.currentText()) as workbook2:
                sheet_names_excel2 = workbook2.sheet_names()
                same_sheet_names = [sheet_name for sheet_name in sheet_names_excel1 if sheet_name in sheet_names_excel2]
                add_sheet_names = [sheet_name for sheet_name in sheet_names_excel2 if sheet_name not in sheet_names_excel1]
                delete_sheet_names = [sheet_name for sheet_name in sheet_names_excel1 if sheet_name not in sheet_names_excel2]

                if len(same_sheet_names) <= 0:
                    QMessageBox.information(self, "warning", u"两个Excel文件无同名sheet")
                    return

                for same_sheet_name in same_sheet_names:
                    diff_sheet_row = 0

                    excel_1, excel_2 = self.create_tab_excel(same_sheet_name)
                    sheet1 = workbook1.sheet_by_name(same_sheet_name)
                    sheet2 = workbook2.sheet_by_name(same_sheet_name)

                    max_nrows = max(sheet1.nrows, sheet2.nrows)
                    max_ncols = max(sheet1.ncols, sheet2.ncols)

                    self.show_origin_excel_sheet(sheet1, excel_1, max_nrows, max_ncols)
                    self.show_origin_excel_sheet(sheet2, excel_2, max_nrows, max_ncols)

                    diff_col, diff_row, diff_cell, diff_sheet = self.create_tab_diff(same_sheet_name)
                    if len(add_sheet_names) > 0:
                        for add_sheet_name in add_sheet_names:
                            self.add_sheet_diff(diff_sheet,diff_sheet_row,add_sheet_name,True)
                            diff_sheet_row += 1
                    if len(delete_sheet_names) > 0:
                        for delete_sheet_name in delete_sheet_names:
                            self.add_sheet_diff(diff_sheet,diff_sheet_row,delete_sheet_name,False)
                            diff_sheet_row += 1

                    diff_row.cellClicked.connect(self.deal_with_row_sym)
                    diff_col.cellClicked.connect(self.deal_with_col_sym)
                    diff_cell.cellClicked.connect(self.deal_with_cell_syn)
                    excel_1.verticalScrollBar().sliderMoved.connect(self.vertical_scroll_bar_syn)
                    excel_2.verticalScrollBar().sliderMoved.connect(self.vertical_scroll_bar_syn)
                    excel_1.horizontalScrollBar().sliderMoved.connect(self.horizontal_scroll_bar_syn)
                    excel_2.horizontalScrollBar().sliderMoved.connect(self.horizontal_scroll_bar_syn)
                    self.deal_with_sheet(same_sheet_name, sheet1, sheet2, diff_col, diff_row, diff_cell, excel_1, excel_2)

    def deal_with_row_sym(self,row, col):
        if col == 1:
            current_excel_1 = self.tabWidgetExcel1.currentWidget().widget(0)
            current_excel_2 = self.tabWidgetExcel1.currentWidget().widget(1)
            current_diff_row = self.tabWidgetDiff.currentWidget().widget(1)

            text_row = current_diff_row.item(row, col)
            show_row = int(text_row.text())-1
            current_excel_1.selectRow(show_row)
            current_excel_1.showRow(show_row)
            current_excel_2.selectRow(show_row)
            current_excel_1.scrollToItem(current_excel_1.item(show_row,0))
            current_excel_2.scrollToItem(current_excel_2.item(show_row,0))

    def deal_with_col_sym(self, row, col):
        if col == 1:
            current_excel_1 = self.tabWidgetExcel1.currentWidget().widget(0)
            current_excel_2 = self.tabWidgetExcel1.currentWidget().widget(1)
            current_diff_col = self.tabWidgetDiff.currentWidget().widget(0)

            text_col = current_diff_col.item(row, col)
            show_col = column_index_from_string(text_col.text()) - 1
            current_excel_1.selectColumn(show_col)
            current_excel_2.selectColumn(show_col)
            current_excel_1.scrollToItem(current_excel_1.item(0, show_col))
            current_excel_2.scrollToItem(current_excel_2.item(0, show_col))

    def deal_with_cell_syn(self, row, col):
        if col == 0:
            current_excel_1 = self.tabWidgetExcel1.currentWidget().widget(0)
            current_excel_2 = self.tabWidgetExcel1.currentWidget().widget(1)
            current_diff_cell = self.tabWidgetDiff.currentWidget().widget(2)

            text = current_diff_cell.item(row, col).text()

            idx1 = text.find(",")
            idx2 = text.find("]")
            new_row = int(text[1:idx1].strip()) - 1
            new_col = column_index_from_string(text[idx1+1:idx2].strip()) - 1
            current_excel_1.setCurrentCell(new_row, new_col)
            current_excel_2.setCurrentCell(new_row, new_col)

    def add_sheet_diff(self,diff_sheet, diff_sheet_row, name, add = True):
        newItem2 = QTableWidgetItem(name)
        if add:
            newItem1 = QTableWidgetItem(u"新增")
            newItem1.setForeground(QColor(0, 0, 255))
            newItem1.setBackground(QColor(0, 255, 255))
            newItem2.setForeground(QColor(0, 0, 255))
            newItem2.setBackground(QColor(0, 255, 255))
        else:
            newItem1 = QTableWidgetItem(u"删除")
            newItem1.setForeground(QColor(255, 0, 0))
            newItem2.setForeground(QColor(255, 0, 0))
        diff_sheet.insertRow(diff_sheet_row)
        diff_sheet.setItem(diff_sheet_row, 0, newItem1)
        diff_sheet.setItem(diff_sheet_row, 1, newItem2)
    # add: True代表新增; False代表删除
    def add_row_diff(self, row, diff_row, diff_row_row, add = True):
        newItem2 = QTableWidgetItem(str(row + 1))
        if add:
            newItem1 = QTableWidgetItem(u"新增")
            newItem1.setForeground(QColor(0, 0, 255))
            newItem1.setBackground(QColor(0, 255, 255))
            newItem2.setForeground(QColor(0, 0, 255))
            newItem2.setBackground(QColor(0, 255, 255))
        else:
            newItem1 = QTableWidgetItem(u"删除")
            newItem1.setForeground(QColor(255, 0, 0))
            newItem2.setForeground(QColor(255, 0, 0))
        diff_row.insertRow(diff_row_row)
        diff_row.setItem(diff_row_row, 0, newItem1)
        diff_row.setItem(diff_row_row, 1, newItem2)

    # add: True代表新增; False代表删除
    def add_col_diff(self, col, diff_col, diff_col_row, add=True):
        newItem2 = QTableWidgetItem(get_column_letter(col + 1))
        if add:
            newItem1 = QTableWidgetItem(u"新增")
            newItem1.setForeground(QColor(0, 0, 255))
            newItem1.setBackground(QColor(0, 255, 255))
            newItem2.setForeground(QColor(0, 0, 255))
            newItem2.setBackground(QColor(0, 255, 255))
        else:
            newItem1 = QTableWidgetItem(u"删除")
            newItem1.setForeground(QColor(255, 0, 0))
            newItem2.setForeground(QColor(255, 0, 0))
        diff_col.insertRow(diff_col_row)
        diff_col.setItem(diff_col_row, 0, newItem1)
        diff_col.setItem(diff_col_row, 1, newItem2)

    def deal_with_sheet(self,sane_sheet_name, sheet1, sheet2, diff_col, diff_row, diff_cell, excel_1, excel_2):
        same = True      # 两个文件比较完成后，是否完全一样
        not_row = []     # 单元格比对时,不需要比较的行
        not_col = []     # 单元格比对时,不需要比较的列

        diff_cell_row = 0
        diff_row_row  = 0
        diff_col_row  = 0

        nrows_sheet1 = sheet1.nrows
        nrows_sheet2 = sheet2.nrows
        ncols_sheet1 = sheet1.ncols
        ncols_sheet2 = sheet2.ncols
        min_nrows = min(nrows_sheet1, nrows_sheet2)
        min_ncols = min(ncols_sheet1, ncols_sheet2)
        max_nrows = max(nrows_sheet1, nrows_sheet2)
        max_ncols = max(ncols_sheet1, ncols_sheet2)

        if nrows_sheet1 == 0 and nrows_sheet2 == 0:
            QMessageBox.information(self, "information", "两个Excel文件中的" + sane_sheet_name + "不存在差异")
            return

        not_empty_row = 0   # 非空的行数
        not_empty_col = 0   # 非空的列数
        if nrows_sheet1 == 0 or nrows_sheet2 == 0:  # 有一个有空表
            if nrows_sheet1 == 0:      # sheet1为空
                for row in range(nrows_sheet2):
                    row_list2 = sheet2.row_values(row)
                    if any(row_list2): # 非空
                        not_empty_row += 1
                        same = False
                for col in range(ncols_sheet2):
                    col_list2 = sheet2.col_values(col)
                    if any(col_list2): # 非空
                        not_empty_col += 1
                        same = False
            if nrows_sheet2 == 0:
                for row in range(nrows_sheet1):
                    row_list1 = sheet1.row_values(row)
                    if any(row_list1): # 非空
                        not_empty_row += 1
                        same = False
                for col in range(ncols_sheet1):
                    col_list1 = sheet1.col_values(col)
                    if any(col_list1): # 非空
                        not_empty_col += 1
                        same = False
        if nrows_sheet1 == 0 or nrows_sheet2 == 0:  # 有一个有空表
            if not_empty_row <= not_empty_col:   # 行为主
                if nrows_sheet1 == 0:      # sheet1为空
                    for row in range(nrows_sheet2):
                        row_list2 = sheet2.row_values(row)
                        if any(row_list2): # 非空
                            self.add_row_diff(row, diff_row, diff_row_row, True)
                            self.set_row_color(excel_1, row, max_ncols, QColor(0, 255, 255))
                            self.set_row_color(excel_2, row, max_ncols, QColor(0, 255, 255))
                            diff_row_row += 1
                            same = False
                if nrows_sheet2 == 0:
                    for row in range(nrows_sheet1):
                        row_list1 = sheet1.row_values(row)
                        if any(row_list1): # 非空
                            self.add_row_diff(row, diff_row, diff_row_row, False)
                            self.set_row_color(excel_1, row, max_ncols, QColor(0, 255, 255))
                            self.set_row_color(excel_2, row, max_ncols, QColor(0, 255, 255))
                            diff_row_row += 1
                            same = False
            else:         # 以列为主
                if nrows_sheet1 == 0:      # sheet1为空
                    for col in range(ncols_sheet2):
                        col_list2 = sheet2.col_values(col)
                        if any(col_list2): # 非空
                            self.add_col_diff(col, diff_col, diff_col_row, True)
                            self.set_col_color(excel_1, max_nrows, col, QColor(0, 255, 255))
                            self.set_col_color(excel_2, max_nrows, col, QColor(0, 255, 255))
                            diff_col_row += 1
                            same = False
                if nrows_sheet2 == 0:
                    for col in range(ncols_sheet1):
                        col_list1 = sheet1.col_values(col)
                        if any(col_list1): # 非空
                            self.add_col_diff(col, diff_col, diff_col_row, False)
                            self.set_col_color(excel_1, max_nrows, col, QColor(255, 0, 0))
                            self.set_col_color(excel_2, max_nrows, col, QColor(255, 0, 0))
                            diff_col_row += 1
                            same = False
        else:                                       # 两个表均不为空
            # 处理行******************************************************
            for row in range(min_nrows):
                row_list1 = sheet1.row_values(row)
                row_list2 = sheet2.row_values(row)
                if (not any(row_list1)) and (any(row_list2)):     # 空和非空, 新增
                    self.add_row_diff(row, diff_row, diff_row_row, True)
                    self.set_row_color(excel_1,row,max_ncols,QColor(0,255,255))
                    self.set_row_color(excel_2,row,max_ncols,QColor(0,255,255))
                    diff_row_row += 1
                    not_row.append(row)
                    same = False
                if  (any(row_list1)) and (not any(row_list2)):     # 删除
                    self.add_row_diff(row, diff_row, diff_row_row, False)
                    self.set_row_color(excel_1, row, max_ncols, QColor("red"))
                    self.set_row_color(excel_2, row, max_ncols, QColor("red"))
                    diff_row_row += 1
                    not_row.append(row)
                    same = False
            if nrows_sheet2 - nrows_sheet1 >= 1:   # 新增
                for row in range(nrows_sheet1, nrows_sheet2):
                    self.add_row_diff(row, diff_row, diff_row_row, True)
                    diff_row_row += 1
                    not_row.append(row)
                    self.set_row_color(excel_1, row, max_ncols, QColor(0,255,255))
                    self.set_row_color(excel_2, row, max_ncols, QColor(0,255,255))
                same = False
            if nrows_sheet1 - nrows_sheet2 >= 1: # 删除
                for row in range(nrows_sheet2, nrows_sheet1):
                    self.add_row_diff(row, diff_row, diff_row_row, False)
                    self.set_row_color(excel_1, row, max_ncols, QColor("red"))
                    self.set_row_color(excel_2, row, max_ncols, QColor("red"))
                    diff_row_row += 1
                    not_row.append(row)
                same = False
            # 处理列******************************************************
            for col in range(min_ncols):
                col_list1 = sheet1.col_values(col)
                col_list2 = sheet2.col_values(col)
                if (not any(col_list1)) and (any(col_list2)):  # 空和非空, 新增
                    self.add_col_diff(col, diff_col, diff_col_row, True)
                    self.set_col_color(excel_1, max_nrows, col, QColor(0,255,255))
                    self.set_col_color(excel_2, max_nrows, col, QColor(0,255,255))
                    diff_col_row += 1
                    not_col.append(col)
                    same = False
                if (any(col_list1)) and (not any(col_list2)):  # 删除
                    self.add_col_diff(col, diff_col, diff_col_row, False)
                    self.set_col_color(excel_1, max_nrows, col, QColor("red"))
                    self.set_col_color(excel_2, max_nrows, col, QColor("red"))
                    diff_col_row += 1
                    not_col.append(col)
                    same = False
            if ncols_sheet2 - ncols_sheet1 >= 1:  # 新增
                for col in range(ncols_sheet1, ncols_sheet2):
                    self.add_col_diff(col, diff_col, diff_col_row, True)
                    self.set_col_color(excel_1, max_nrows, col, QColor(0, 255, 255))
                    self.set_col_color(excel_2, max_nrows, col, QColor(0, 255, 255))
                    diff_col_row += 1
                    not_col.append(col)
                same = False
            if ncols_sheet1 - ncols_sheet2 >= 1:  # 删除
                for col in range(ncols_sheet2, ncols_sheet1):
                    self.add_col_diff(col, diff_col, diff_col_row, False)
                    self.set_col_color(excel_1, max_nrows, col, QColor("red"))
                    self.set_col_color(excel_2, max_nrows, col, QColor("red"))
                    diff_col_row += 1
                    not_col.append(col)
                same = False
            # 处理单元格******************************************************
            for row in range(min_nrows):
                if row not in not_row:
                    for col in range(min_ncols):
                        if col not in not_col:

                            # sheet1_celll = sheet1.cell_value(row, col)
                            # sheet1_cell2 = sheet2.cell_value(row, col)
                            sheet1_celll = self.to_str(sheet1.cell_value(row, col))
                            sheet1_cell2 = self.to_str(sheet2.cell_value(row, col))

                            if sheet1_celll != sheet1_cell2:
                                same = False
                                self.set_cell_color(excel_1, row, col, QColor("yellow"))
                                self.set_cell_color(excel_2, row, col, QColor("yellow"))

                                col = get_column_letter(col + 1)
                                text = "[%d, %s]、[%d, %s]" % (row + 1, col, row + 1, col)


                                newItem1 = QTableWidgetItem(sheet1_celll)
                                newItem2 = QTableWidgetItem(sheet1_cell2)
                                newItem3 = QTableWidgetItem(text)
                                newItem3.setForeground(QColor(0, 0, 255))  # 字体设置成蓝色
                                newItem3.setBackground(QColor("yellow"))   # 背景设置成黄色
                                diff_cell.insertRow(diff_cell_row)
                                diff_cell.setItem(diff_cell_row, 0, newItem3)
                                diff_cell.setItem(diff_cell_row, 1, newItem1)
                                diff_cell.setItem(diff_cell_row, 2, newItem2)
                                diff_cell_row += 1
        if same:
            QMessageBox.information(self, "information", "两个Excel文件中的"+sane_sheet_name+"不存在差异")

    def to_str(self, text):
        if isinstance(text, (float, int)):
            text = str(int(text))
        return text

    def vertical_scroll_bar_syn(self,value):
        current_excel_1 = self.tabWidgetExcel1.currentWidget().widget(0)
        current_excel_2 = self.tabWidgetExcel1.currentWidget().widget(1)
        current_excel_1.verticalScrollBar().setValue(value)
        current_excel_2.verticalScrollBar().setValue(value)

    def horizontal_scroll_bar_syn(self,value):
        current_excel_1 = self.tabWidgetExcel1.currentWidget().widget(0)
        current_excel_2 = self.tabWidgetExcel1.currentWidget().widget(1)
        current_excel_1.horizontalScrollBar().setValue(value)
        current_excel_2.horizontalScrollBar().setValue(value)

    # 创建子窗口, 这个可以设计成一个类
    def create_tab_excel(self, name):
        excel_1 = QTableWidget()
        excel_2 = QTableWidget()

        splitter = QSplitter()
        splitter.addWidget(excel_1)
        splitter.addWidget(excel_2)
        splitter.setOrientation(Qt.Horizontal)

        self.tabWidgetExcel1.addTab(splitter, name)
        return excel_1, excel_2

    # 创建子窗口, 这个可以设计成一个类
    def create_tab_diff(self, name):
        diff_sheet = QTableWidget()  # 显示sheet改动
        diff_sheet.setColumnCount(2)
        diff_sheet.setHorizontalHeaderLabels(['改动', 'sheet名'])
        diff_col = QTableWidget()  # 显示列改动
        diff_col.setColumnCount(2)
        diff_col.setHorizontalHeaderLabels(['改动', '列号'])
        diff_row = QTableWidget()  # 显示行改动
        diff_row.setColumnCount(2)
        diff_row.setHorizontalHeaderLabels(['改动', '行号'])
        diff_cell = QTableWidget()  # 显示单元格改动
        diff_cell.setColumnCount(3)
        diff_cell.setHorizontalHeaderLabels(['坐标', '旧值', '新值'])

        splitter = QSplitter()
        splitter.addWidget(diff_col)
        splitter.addWidget(diff_row)
        splitter.addWidget(diff_cell)
        splitter.addWidget(diff_sheet)
        self.tabWidgetDiff.addTab(splitter, name+"-diff")

        return diff_col, diff_row, diff_cell, diff_sheet

    # 相关的窗口进行连动
    def update_tab(self,num):
        self.tabWidgetExcel1.setCurrentWidget(self.tabWidgetExcel1.widget(num))
        self.tabWidgetDiff.setCurrentWidget(self.tabWidgetDiff.widget(num))

    def set_cell_color(self, tableWidget, row, col, color):
        item = tableWidget.item(row, col)
        if item is None:
            item = QTableWidgetItem()
            item.setBackground(color)
            tableWidget.setItem(row, col, item)
        else:
            item.setBackground(color)

    def set_row_color(self, tableWidget, row, max_col, color):
        for col in range(max_col):
            self.set_cell_color(tableWidget, row, col, color)

    def set_col_color(self, tableWidget, max_row, col, color):
        for row in range(max_row):
            self.set_cell_color(tableWidget, row, col, color)


def num_converted_into_letters(num):
    # 如将数字1，2，3...转化成A，B，C... 27转化成AA....
    result = []
    for i in range(num):
        result.append(get_column_letter(i+1))
    return result

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon("./images/Sync.png"))
    form = MainWindow()
    form.show()
    sys.exit(app.exec())